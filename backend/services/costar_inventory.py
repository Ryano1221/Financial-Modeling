from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

COLUMN_MAP = {
    "longitude": "Longitude",
    "latitude": "Latitude",
    "submarket": "Submarket Name",
    "property_type": "Property Type",
    "name": "Property Name",
    "address": "Property Address",
    "city": "City",
    "state": "State",
    "market": "Market Name",
    "building_status": "Building Status",
    "year_built": "Year Built",
    "year_renovated": "Year Renovated",
    "building_class": "Building Class",
    "stories": "Number Of Stories",
    "core_factor": "Core Factor",
    "rba": "RBA",
    "amenities": "Amenities",
    "typical_floor_size": "Typical Floor Size",
    "operating_expenses": "Building Operating Expenses",
    "parking_ratio": "Parking Ratio",
    "owner_name": "Owner Name",
    "owner_phone": "Owner Phone",
    "property_manager_name": "Property Manager Name",
    "property_manager_phone": "Property Manager Phone",
    "leasing_company_name": "Leasing Company Name",
    "leasing_company_contact": "Leasing Company Contact",
    "leasing_company_phone": "Leasing Company Phone",
    "property_id": "PropertyID",
}


def _as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_number(value: object) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = _as_text(value).replace(",", "")
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None
    if number.is_integer():
        return int(number)
    return round(number, 6)


def _normalized_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for index, value in enumerate(header_row):
        text = _as_text(value)
        if text:
            lookup[text] = index
    return lookup


def _row_value(row: tuple[Any, ...], header_map: dict[str, int], column_name: str) -> object:
    index = header_map.get(column_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _dedupe_key(record: dict[str, Any]) -> str:
    property_id = _as_text(record.get("propertyId"))
    if property_id:
        return f"property:{property_id}"
    return "|".join(
        [
            _as_text(record.get("name")).lower(),
            _as_text(record.get("address")).lower(),
            _as_text(record.get("submarket")).lower(),
        ]
    )


def _build_record(row: tuple[Any, ...], header_map: dict[str, int], source_note: str) -> dict[str, Any]:
    property_id = _as_text(_row_value(row, header_map, COLUMN_MAP["property_id"]))
    name = _as_text(_row_value(row, header_map, COLUMN_MAP["name"]))
    fallback_slug = name.lower().replace(" ", "_") or "building"
    city = _as_text(_row_value(row, header_map, COLUMN_MAP["city"]))
    state = _as_text(_row_value(row, header_map, COLUMN_MAP["state"]))
    market = _as_text(_row_value(row, header_map, COLUMN_MAP["market"]))
    submarket = _as_text(_row_value(row, header_map, COLUMN_MAP["submarket"]))
    property_type = _as_text(_row_value(row, header_map, COLUMN_MAP["property_type"]))
    return {
        "id": f"costar_{property_id or fallback_slug}",
        "clientId": "market_inventory_shared",
        "name": name,
        "address": _as_text(_row_value(row, header_map, COLUMN_MAP["address"])),
        "city": city,
        "state": state,
        "market": market,
        "submarket": submarket,
        "ownerName": _as_text(_row_value(row, header_map, COLUMN_MAP["owner_name"])),
        "propertyType": property_type,
        "totalRSF": _as_number(_row_value(row, header_map, COLUMN_MAP["rba"])) or 0,
        "notes": "Imported from CoStar office inventory export.",
        "buildingClass": _as_text(_row_value(row, header_map, COLUMN_MAP["building_class"])),
        "buildingStatus": _as_text(_row_value(row, header_map, COLUMN_MAP["building_status"])),
        "yearBuilt": _as_number(_row_value(row, header_map, COLUMN_MAP["year_built"])),
        "yearRenovated": _as_number(_row_value(row, header_map, COLUMN_MAP["year_renovated"])),
        "numberOfStories": _as_number(_row_value(row, header_map, COLUMN_MAP["stories"])),
        "coreFactor": _as_number(_row_value(row, header_map, COLUMN_MAP["core_factor"])),
        "typicalFloorSize": _as_number(_row_value(row, header_map, COLUMN_MAP["typical_floor_size"])),
        "parkingRatio": _as_number(_row_value(row, header_map, COLUMN_MAP["parking_ratio"])),
        "operatingExpenses": _as_text(_row_value(row, header_map, COLUMN_MAP["operating_expenses"])),
        "amenities": _as_text(_row_value(row, header_map, COLUMN_MAP["amenities"])),
        "propertyId": property_id,
        "ownerPhone": _as_text(_row_value(row, header_map, COLUMN_MAP["owner_phone"])),
        "propertyManagerName": _as_text(_row_value(row, header_map, COLUMN_MAP["property_manager_name"])),
        "propertyManagerPhone": _as_text(_row_value(row, header_map, COLUMN_MAP["property_manager_phone"])),
        "leasingCompanyName": _as_text(_row_value(row, header_map, COLUMN_MAP["leasing_company_name"])),
        "leasingCompanyContact": _as_text(_row_value(row, header_map, COLUMN_MAP["leasing_company_contact"])),
        "leasingCompanyPhone": _as_text(_row_value(row, header_map, COLUMN_MAP["leasing_company_phone"])),
        "latitude": _as_number(_row_value(row, header_map, COLUMN_MAP["latitude"])),
        "longitude": _as_number(_row_value(row, header_map, COLUMN_MAP["longitude"])),
        "source": source_note,
      }


def parse_costar_inventory_workbook(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []
    header_map = _normalized_header_map(header_row)
    required = [COLUMN_MAP["name"], COLUMN_MAP["address"], COLUMN_MAP["property_type"]]
    missing = [column for column in required if column not in header_map]
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(f"Missing required CoStar columns: {missing_text}")

    source_note = f"CoStar import uploaded via account settings: {filename}"
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        property_type = _as_text(_row_value(row, header_map, COLUMN_MAP["property_type"]))
        if "office" not in property_type.lower():
            continue
        record = _build_record(row, header_map, source_note)
        if not _as_text(record.get("name")) and not _as_text(record.get("address")):
            continue
        key = _dedupe_key(record)
        if not key or key in seen:
            continue
        seen.add(key)
        records.append(record)
    records.sort(key=lambda item: (
        _as_text(item.get("market")).lower(),
        _as_text(item.get("submarket")).lower(),
        _as_text(item.get("name")).lower(),
        _as_text(item.get("address")).lower(),
    ))
    return records


def merge_market_inventory(existing_records: list[dict[str, Any]], imported_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for record in existing_records:
        merged[_dedupe_key(record)] = dict(record)
    for record in imported_records:
        key = _dedupe_key(record)
        prior = merged.get(key, {})
        merged[key] = {**prior, **record}
    return sorted(
        merged.values(),
        key=lambda item: (
            _as_text(item.get("market")).lower(),
            _as_text(item.get("submarket")).lower(),
            _as_text(item.get("name")).lower(),
            _as_text(item.get("address")).lower(),
        ),
    )


def build_market_inventory_envelope(
    records: list[dict[str, Any]],
    *,
    filename: str,
    imported_by_user_id: str,
    imported_by_email: str,
    previous_count: int,
) -> dict[str, Any]:
    updated_at = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return {
        "version": 1,
        "updated_at": updated_at,
        "source": "costar_excel_import",
        "records": records,
        "summary": {
            "count": len(records),
            "import_filename": filename,
            "imported_by_user_id": imported_by_user_id,
            "imported_by_email": imported_by_email,
            "previous_count": previous_count,
            "delta": len(records) - previous_count,
        },
    }
